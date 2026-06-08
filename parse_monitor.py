"""
parse_monitor.py
Extrae snapshots de capacidad de generación del fichero consolidado de monitorización
'Monitor_Capacidad_Red_INTEGRADO_v4.xlsx'.

Estructura del Excel:
  - Hojas DSO Xxx:  Una fila por SET (todas las distribuidoras integradas)
      fecha   → row3 col[1]  (datetime)
      col[1]  → Nombre DSO
      col[3]  → Provincia   col[4] → Municipio   col[5] → Nombre Subestación
      col[9]  → Tensión (kV)
      col[21] → Capacidad disponible (MW) [generación]
      col[26] → Capacidad admitida y no resuelta (MW) [en estudio]
      disponible_neta = col[21] - col[26]

  - Hojas REE Xxx:  Un nudo por fila
      fecha   → row3 col[1]  (datetime)
      col[1]  → Nombre y tensión del nudo
      col[3]  → Comunidad Autónoma
      col[40] → Solicitada en curso MPE (total)
      col[57] → Capacidad disponible MPE RdT [MW]
      disponible_neta = col[57] - col[40]

Salida: references/monitor_capacidad.csv
  (mismo schema que dso_capacidad_historial.csv)

Uso:
  python parse_monitor.py
  python parse_monitor.py --summary
"""
import re, sys, argparse
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl --break-system-packages")
try:
    import pandas as pd
except ImportError:
    sys.exit("pip install pandas --break-system-packages")

BASE = Path(__file__).parent
XLSX = BASE / "Monitor_Capacidad_Red_INTEGRADO_v4.xlsx"
OUT  = BASE / "references" / "monitor_capacidad.csv"

_TENSION_RE = re.compile(r'\b(\d{2,4})\s*$')

def to_float(v):
    if v is None or str(v).strip() in ('', 'N/A', 'N/D', '-', 'NA', 'Not found', '#N/A'):
        return None
    try:
        return float(str(v).replace(',', '.'))
    except (ValueError, TypeError):
        return None

def extract_tension(nombre):
    if not nombre: return None
    m = _TENSION_RE.search(str(nombre).strip())
    return int(m.group(1)) if m else None

def strip_tension(nombre):
    if not nombre: return ''
    return _TENSION_RE.sub('', str(nombre).strip()).strip()

def get_fecha(rows, col=1):
    """Extrae fecha de row3 (index 2)."""
    v = rows[2][col] if len(rows) > 2 and len(rows[2]) > col else None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if v and str(v).strip():
        try:
            return pd.to_datetime(str(v)).strftime('%Y-%m-%d')
        except Exception:
            pass
    return None

# ── DSO processing ─────────────────────────────────────────────────────────────
def process_dso_sheet(rows, fecha):
    """
    Header at index 4 (row 5). Data from index 5.
    col[1] DSO, col[3] prov, col[4] mun, col[5] nombre SET, col[9] kV,
    col[21] disponible, col[26] admitida_no_resuelta
    """
    records = []
    for row in rows[5:]:
        if not row or len(row) < 27:
            continue
        nombre = row[5]
        if not nombre or str(nombre).strip() in ('', 'Nombre Subestación'):
            continue
        nombre = str(nombre).strip()
        if nombre.startswith('='):
            continue

        dso      = str(row[1] or '').strip()
        # Fallback DSO: en Oct-2025 las filas UFD tienen col[1] vacío
        # pero col[2] contiene el código de gestor de red (R1-XXX)
        if not dso:
            gestor = str(row[2] or '').strip()
            if gestor.startswith('R1-'):
                dso = 'UFD'
        if not dso:
            continue   # Fila sin distribuidora identificable
        prov     = str(row[3] or '').strip()
        mun      = str(row[4] or '').strip()
        tension  = to_float(row[9])
        disp     = to_float(row[21])
        est      = to_float(row[26]) or 0.0

        if disp is None:
            continue

        records.append({
            'fecha':           fecha,
            'distribuidora':   dso,
            'provincia':       prov,
            'municipio':       mun,
            'subestacion':     nombre,
            'tension_kv':      tension,
            'disponible_neta': round(disp - est, 4),
        })
    return records

# ── REE processing ─────────────────────────────────────────────────────────────
def process_ree_sheet(rows, fecha):
    """
    Header at index 6 (row 7). Data from index 7 (row 8).
    col[1] nudo+tensión, col[3] CCAA,
    col[40] solicitada MPE, col[57] disponible MPE RdT
    """
    COL_NUDO = 1
    COL_CCAA = 3
    COL_SOL  = 40
    COL_DISP = 57

    records = []
    for row in rows[7:]:
        if not row or len(row) <= COL_DISP:
            continue
        nudo_raw = row[COL_NUDO]
        if not nudo_raw or str(nudo_raw).strip() in ('', 'Nombre y tensión del nudo'):
            continue
        nudo_raw = str(nudo_raw).strip()
        if nudo_raw.startswith('='):
            continue

        disp = to_float(row[COL_DISP])
        if disp is None:
            continue
        sol  = to_float(row[COL_SOL]) or 0.0

        ccaa    = str(row[COL_CCAA] or '').strip() if len(row) > COL_CCAA else ''
        tension = extract_tension(nudo_raw)
        nombre  = strip_tension(nudo_raw)

        records.append({
            'fecha':           fecha,
            'distribuidora':   'REE',
            'provincia':       '',
            'municipio':       '',
            'subestacion':     nombre or nudo_raw,
            'tension_kv':      tension,
            'disponible_neta': round(disp - sol, 4),
            'ccaa':            ccaa,
        })
    return records

# ── main ──────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument('--summary', action='store_true')
args = parser.parse_args()

if not XLSX.exists():
    sys.exit(f"No se encuentra {XLSX}")

print(f"Leyendo {XLSX.name} …")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

all_records = []
seen_dso_fechas = set()
seen_ree_fechas = set()

DSO_SHEETS = [s for s in wb.sheetnames
              if s.startswith('DSO ') and s not in ('DSO Anterior',)]
REE_SHEETS = [s for s in wb.sheetnames
              if s.startswith('REE ') and s not in ('REE Anterior',)]

for sname in DSO_SHEETS:
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    fecha = get_fecha(rows)
    if not fecha:
        print(f"  SKIP {sname!r}: fecha no detectada")
        continue
    if fecha in seen_dso_fechas:
        print(f"  SKIP {sname!r}: fecha {fecha} ya procesada (duplicado)")
        continue
    seen_dso_fechas.add(fecha)

    records = process_dso_sheet(rows, fecha)
    all_records.extend(records)
    dsos = sorted(set(r['distribuidora'] for r in records))
    print(f"  {fecha}  ({sname!r:<15})  DSO  →  {len(records):>5} SETs  [{', '.join(dsos)}]")

for sname in REE_SHEETS:
    ws = wb[sname]
    rows = list(ws.iter_rows(values_only=True))
    fecha = get_fecha(rows)
    if not fecha:
        print(f"  SKIP {sname!r}: fecha no detectada")
        continue
    if fecha in seen_ree_fechas:
        print(f"  SKIP {sname!r}: fecha {fecha} ya procesada (duplicado)")
        continue
    seen_ree_fechas.add(fecha)

    records = process_ree_sheet(rows, fecha)
    all_records.extend(records)
    print(f"  {fecha}  ({sname!r:<15})  REE  →  {len(records):>5} nudos")

wb.close()

if not all_records:
    sys.exit("No se encontraron registros.")

df = pd.DataFrame(all_records)
df['fecha'] = pd.to_datetime(df['fecha'])
df = df.sort_values(['distribuidora', 'subestacion', 'tension_kv', 'fecha'])

print(f"\n--- RESUMEN ---")
print(f"Total registros  : {len(df):,}")
print(f"Distribuidoras   : {sorted(df['distribuidora'].unique().tolist())}")
print(f"Snapshots DSO    : {len(seen_dso_fechas)}  REE: {len(seen_ree_fechas)}")
print(f"Rango fechas     : {df['fecha'].min().date()} → {df['fecha'].max().date()}")
print(f"SETs únicos      : {df['subestacion'].nunique():,}")

if args.summary:
    sys.exit(0)

df['fecha'] = df['fecha'].dt.strftime('%Y-%m-%d')
df.to_csv(OUT, index=False, encoding='utf-8-sig')
print(f"Guardado en      : {OUT}")
