#!/usr/bin/env python3
"""
parse_capacity_excel_history.py v2
Parsea Excel UFD/iDE/REE con hojas de fecha y extiende sets_history.json.
Auto-detecta la fila de cabecera y posición de columnas.
"""
import json, re, unicodedata, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl --break-system-packages")

BASE = Path(__file__).parent
HIST = BASE / "sets_history.json"
SETS = BASE / "sets_capacity.json"

MESES = {
    "enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
    "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12
}

def norm(s):
    s = str(s or "").upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def to_float(v):
    if v is None or str(v).strip() in ("","N/A","#N/A","None"): return None
    try: return float(str(v).replace(",","."))
    except: return None

def parse_sheet_date(name):
    parts = name.strip().lower().split()
    if len(parts)==3:
        m = MESES.get(parts[1][:3]) or MESES.get(parts[1])
        if m:
            try: return f"{int(parts[2]):04d}-{m:02d}"
            except: pass
    return None

# ── Load bases ────────────────────────────────────────────────────────────────
print("Loading sets_history.json…")
with open(HIST) as f: history = json.load(f)
print("Loading sets_capacity.json…")
with open(SETS) as f: capacity = json.load(f)

# Build norm index
norm_to_key = {}
for k in capacity:
    nk = norm(k)
    norm_to_key[nk] = k
    base = re.sub(r"\s+\d+$","",nk)
    norm_to_key.setdefault(base, k)

def match_set(nombre, kv=None):
    nname = norm(str(nombre))
    if nname in norm_to_key: return norm_to_key[nname]
    if kv is not None:
        try:
            kv_int = str(int(float(kv)))
            cand = nname + " " + kv_int
            if cand in norm_to_key: return norm_to_key[cand]
        except: pass
    words = [w for w in nname.split() if len(w)>=4]
    if not words: return None
    best, best_s = None, 0
    for kn, kv2 in norm_to_key.items():
        hits = sum(1 for w in words if w in kn)
        s = hits/len(words)
        if s>=0.75 and s>best_s: best_s=s; best=kv2
    return best

added = skipped = 0

def add_entry(key, date_str, cap_gen=None, cap_gen_ocup=None):
    global added, skipped
    if key not in history: history[key] = []
    if any(e.get("date")==date_str for e in history[key]):
        skipped+=1; return
    e = {"date": date_str}
    if cap_gen is not None: e["cap_gen"] = round(cap_gen, 2)
    if cap_gen_ocup is not None: e["cap_gen_ocup"] = round(cap_gen_ocup, 2)
    history[key].append(e)
    added+=1

def find_col(headers, *keywords):
    """Find column index by keyword match (case-insensitive)."""
    for kw in keywords:
        kw_l = kw.lower()
        for i,h in enumerate(headers):
            if h and kw_l in str(h).lower():
                return i
    return None

def parse_sheet_auto(ws, date_str, src, header_row_max=8):
    """Auto-detect header row, map columns, parse data."""
    all_rows = list(ws.iter_rows(values_only=True))
    # Find header row
    hdr_idx = None
    hdr = None
    for i,row in enumerate(all_rows[:header_row_max]):
        row_text = " ".join(str(c or "").lower() for c in row if c)
        if any(k in row_text for k in ["nombre","denominación","denominacion","nombre y tensión","nombre y tension"]):
            hdr_idx = i
            hdr = [str(c or "") for c in row]
            break
    if hdr_idx is None: return 0

    data_rows = all_rows[hdr_idx+1:]
    n = 0

    if src == "UFD":
        col_nombre = find_col(hdr, "nombre")
        col_kv     = find_col(hdr, "nivel de tensión", "nivel de tension", "nivel")
        col_pos    = find_col(hdr, "posición del parque", "posicion del parque", "posición","posicion")
        col_disp   = find_col(hdr, "capacidad disponible")
        col_ocup   = find_col(hdr, "capacidad ocupada","ocupada")
        if col_nombre is None or col_disp is None: return 0
        for row in data_rows:
            if not row or len(row) <= max(c for c in [col_nombre,col_kv,col_pos,col_disp,col_ocup] if c is not None): continue
            nombre = row[col_nombre]
            if not nombre: continue
            pos = str(row[col_pos] or "") if col_pos is not None else ""
            if "total" not in pos.lower(): continue
            kv   = row[col_kv] if col_kv is not None else None
            disp = to_float(row[col_disp])
            ocup = to_float(row[col_ocup]) if col_ocup is not None else None
            key  = match_set(str(nombre), kv)
            if not key: continue
            add_entry(key, date_str, cap_gen=disp, cap_gen_ocup=ocup)
            n+=1

    elif src == "iDE":
        col_nombre = find_col(hdr, "denominación del punto","denominacion del punto","nombre","denominación","denominacion")
        col_kv     = find_col(hdr, "nivel de tensión","nivel de tension","nivel")
        col_disp   = find_col(hdr, "capacidad de acceso disponible","disponible")
        col_ocup   = find_col(hdr, "capacidad de acceso ocupada","ocupada")
        if col_nombre is None or col_disp is None: return 0
        for row in data_rows:
            if not row: continue
            nombre = row[col_nombre] if col_nombre < len(row) else None
            if not nombre: continue
            kv   = row[col_kv] if col_kv is not None and col_kv < len(row) else None
            disp = to_float(row[col_disp]) if col_disp < len(row) else None
            ocup = to_float(row[col_ocup]) if col_ocup is not None and col_ocup < len(row) else None
            key  = match_set(str(nombre), kv)
            if not key: continue
            add_entry(key, date_str, cap_gen=disp, cap_gen_ocup=ocup)
            n+=1

    elif src == "REE":
        # Col0 = "Nombre y tensión del nudo", col to find = "Capacidad de acceso nodal" or "Margen no ocupado"
        # Try to find static criterion column (col4 or col6)
        col_nudo   = 0  # always first
        # Find best capacity column: prefer "margen no ocupado" (available), fallback to "capacidad de acceso nodal"
        col_margen = find_col(hdr, "margen no ocupado")
        col_nodal  = find_col(hdr, "capacidad de acceso nodal","capacidad nodal")
        col_use    = col_margen if col_margen is not None else col_nodal
        if col_use is None: return 0
        for row in data_rows:
            if not row or not row[0]: continue
            nudo = str(row[0]).strip()
            if not nudo or nudo.lower() in ("nombre y tensión del nudo","nombre y tension"): continue
            val = to_float(row[col_use]) if col_use < len(row) else None
            # Parse "ABADIANO 220"
            m = re.match(r'^(.+?)\s+(\d+)\s*$', nudo)
            if m:
                nombre, kv = m.group(1).strip(), m.group(2)
                key = match_set(nombre, kv)
            else:
                key = match_set(nudo)
            if not key: continue
            add_entry(key, date_str, cap_gen=val)
            n+=1
    return n

# ── UFD ───────────────────────────────────────────────────────────────────────
ufd = BASE/"references"/"Mapas_Capacidad_AyC-UFD.xlsx"
if ufd.exists():
    print(f"\nParsing UFD ({ufd.name})")
    wb = openpyxl.load_workbook(ufd, read_only=True, data_only=True)
    for sname in wb.sheetnames:
        ds = parse_sheet_date(sname)
        if not ds: continue
        n = parse_sheet_auto(wb[sname], ds, "UFD")
        print(f"  {sname}: {n} entries")
    wb.close()

# ── iDE ───────────────────────────────────────────────────────────────────────
ide = BASE/"references"/"Mapas_Capacidad_AyC-iDE.xlsx"
if ide.exists():
    print(f"\nParsing iDE ({ide.name})")
    wb = openpyxl.load_workbook(ide, read_only=True, data_only=True)
    skip = {"Resumen Nov","Resumen Dic","Resumen Ene","Demanda_BaseDatos","Generacion_BaseDatos"}
    for sname in wb.sheetnames:
        if sname in skip: continue
        ds = parse_sheet_date(sname)
        if not ds: continue
        n = parse_sheet_auto(wb[sname], ds, "iDE")
        print(f"  {sname}: {n} entries")
    wb.close()

# ── REE ───────────────────────────────────────────────────────────────────────
ree = BASE/"references"/"Mapas_Capacidad_AyC-REE.xlsx"
if ree.exists():
    print(f"\nParsing REE ({ree.name})")
    wb = openpyxl.load_workbook(ree, read_only=True, data_only=True)
    skip = {"Diccionario nudos capacidad zon","RdD Aceptabilidad","Diccionario"}
    for sname in wb.sheetnames:
        if sname in skip: continue
        ds = parse_sheet_date(sname)
        if not ds: continue
        n = parse_sheet_auto(wb[sname], ds, "REE")
        print(f"  {sname}: {n} entries")
    wb.close()

# Sort + save
for k in history: history[k].sort(key=lambda e: e.get("date",""))
sets_with_hist = sum(1 for v in history.values() if v)
print(f"\nAdded: {added}  Skipped (dup): {skipped}")
print(f"SETs with history: {sets_with_hist}")

with open(HIST,"w") as f:
    json.dump(history, f, ensure_ascii=False, separators=(",",":"))
print(f"Written: {HIST}")
