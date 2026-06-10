"""
parse_ree.py
Extrae capacidad MPE de REE desde 'references/Mapas_Capacidad_AyC-REE.xlsx'.

Columnas extraídas:
  disponible_neta = disp_MPE_RdT - solicitada_MPE
  cap_gen_ocup    = otorgada_MPE
  cap_gen_tram    = solicitada_MPE

Formato antiguo (≤2024):  nombre=A(0),  ccaa=B(1),  disp=AB(27), otorg=S(18)/AH(33), sol=U(20)
Formato nuevo  (≥2024):   nombre=C(2),  ccaa=D(3),  disp=BC(54), otorg=AA(26),       sol=AH(33)

Uso:
  python parse_ree.py
  python parse_ree.py --all-dates
  python parse_ree.py --summary
"""
import re, sys, argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl --break-system-packages")
try:
    import pandas as pd
except ImportError:
    sys.exit("pip install pandas --break-system-packages")

BASE = Path(__file__).parent
XLSX = BASE / "references" / "Mapas_Capacidad_AyC-REE.xlsx"
OUT  = BASE / "references" / "ree_capacidad.csv"

MESES = {
    'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
    'julio':7,'agosto':8,'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12
}
_TENSION_RE = re.compile(r'\b(\d{2,4})\s*$')


def parse_sheet_date(name):
    m = re.match(r'(\d{1,2})\s+(\w+)\s+(\d{4})', name.strip())
    if not m:
        return None
    day, mes, year = int(m.group(1)), m.group(2).lower(), int(m.group(3))
    return f'{year}-{MESES[mes]:02d}-{day:02d}' if mes in MESES else None


def detect_columns(header_row):
    """
    Detecta índices de columnas en la fila de cabecera.
    Devuelve dict con: nombre, ccaa, disp, sol, otorg  (otorg puede ser None)
    """
    col_nombre = col_ccaa = col_disp = col_sol = col_otorg = None
    sol_candidates = []
    otorg_candidates = []

    for j, v in enumerate(header_row):
        if not v:
            continue
        sl = str(v).strip().lower().replace('\n', ' ')

        if col_nombre is None and 'nombre y tens' in sl:
            col_nombre = j

        if col_ccaa is None and 'comunidad aut' in sl:
            col_ccaa = j

        # Disponible MPE RdT
        if col_disp is None and 'disponible para mpe rdt' in sl:
            col_disp = j
        if col_disp is None and re.search(r'^mpe rdt\s*\[mw\]$', sl):
            col_disp = j

        # Solicitada MPE (en trámite)
        if 'solicitad' in sl:
            if sl.endswith(' mpe'):               # formato nuevo: "...pendiente resolver MPE"
                col_sol = j
            elif re.search(r'solicitada.* mpe$', sl):
                col_sol = j
            else:
                sol_candidates.append(j)

        # Otorgada MPE (ocupada)
        if 'otorgad' in sl:
            if re.search(r'otorgada mpe$', sl):   # "Capacidad de acceso otorgada MPE" (preferente, formato detallado)
                col_otorg = j
            elif col_otorg is None and re.search(r'otorgada para mpe\s*\[mw\]', sl):  # formato antiguo resumen (solo si no hay match preferente)
                col_otorg = j
            elif col_otorg is None:
                otorg_candidates.append(j)

    # Fallbacks
    if col_sol is None and sol_candidates:
        col_sol = sol_candidates[-1]
    if col_otorg is None and otorg_candidates:
        # Preferir "para MPE" sobre otros
        col_otorg = otorg_candidates[-1]

    if col_nombre is None or col_disp is None or col_sol is None:
        return None

    return {'nombre': col_nombre, 'ccaa': col_ccaa,
            'disp': col_disp, 'sol': col_sol, 'otorg': col_otorg}


def to_float(v):
    if v is None or str(v).strip() in ('', 'N/A', 'N/D', '-', 'NA'):
        return None
    try:
        return float(str(v).replace(',', '.'))
    except (ValueError, TypeError):
        return None

def parse_tension(nombre):
    m = _TENSION_RE.search(str(nombre).strip())
    return int(m.group(1)) if m else None

def strip_tension(nombre):
    return _TENSION_RE.sub('', str(nombre).strip()).strip()


def process_sheet(ws, fecha, cols):
    records = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        nombre_raw = row[cols['nombre']] if cols['nombre'] < len(row) else None
        if not nombre_raw:
            continue
        nombre_raw = str(nombre_raw).strip()
        if not nombre_raw or nombre_raw in ('For Formula ROW',):
            continue
        if nombre_raw.upper().startswith(('NOMBRE', 'TOTAL')):
            continue

        disp_val = to_float(row[cols['disp']] if cols['disp'] < len(row) else None)
        if disp_val is None:
            continue

        sol_val   = to_float(row[cols['sol']]   if cols['sol']   < len(row) else None) or 0.0
        otorg_val = to_float(row[cols['otorg']]  if cols['otorg'] is not None and cols['otorg'] < len(row) else None)

        ccaa = str(row[cols['ccaa']] or '').strip() if cols['ccaa'] is not None and cols['ccaa'] < len(row) else ''

        records.append({
            'fecha':           fecha,
            'distribuidora':   'REE',
            'provincia':       '',
            'municipio':       '',
            'subestacion':     strip_tension(nombre_raw) or nombre_raw,
            'tension_kv':      parse_tension(nombre_raw),
            'disponible_bruta': round(disp_val, 4),
            'disponible_neta': round(disp_val - sol_val, 4),
            'cap_gen_ocup':    otorg_val,
            'cap_gen_tram':    sol_val if sol_val != 0 else None,
            'ccaa':            ccaa,
        })
    return records


# ── main ──────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument('--all-dates', action='store_true')
parser.add_argument('--summary',   action='store_true')
args = parser.parse_args()

MIN_DATE = '0000-01-01' if args.all_dates else '2023-12-01'

if not XLSX.exists():
    sys.exit(f"No se encuentra {XLSX}")

print(f"Leyendo {XLSX.name} …")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

all_records, skipped = [], []

for sname in wb.sheetnames:
    fecha = parse_sheet_date(sname)
    if not fecha or fecha < MIN_DATE:
        if fecha:
            skipped.append((sname, f'antes de {MIN_DATE}'))
        continue

    ws     = wb[sname]
    header = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    cols   = detect_columns(header)

    if cols is None:
        skipped.append((sname, 'columnas no detectadas'))
        continue

    records = process_sheet(ws, fecha, cols)
    all_records.extend(records)
    otorg_str = f", otorg={cols['otorg']}" if cols['otorg'] is not None else ""
    print(f"  {fecha}  ({sname!r:<22})  →  {len(records):>4} nudos  "
          f"[disp={cols['disp']}, sol={cols['sol']}{otorg_str}]")

wb.close()

if skipped:
    print(f"\n  Omitidas ({len(skipped)}):")
    for s, r in skipped[:10]:
        print(f"    {s!r}: {r}")

if not all_records:
    sys.exit("\nNo se encontraron registros.")

df = pd.DataFrame(all_records)
df['fecha'] = pd.to_datetime(df['fecha'])
df = df.sort_values(['subestacion', 'tension_kv', 'fecha'])

print(f"\n--- RESUMEN ---")
print(f"Total registros  : {len(df):,}")
print(f"Snapshots        : {df['fecha'].nunique()} ({df['fecha'].min().date()} → {df['fecha'].max().date()})")
print(f"Nudos únicos     : {df['subestacion'].nunique():,}")
print(f"Con cap_gen_ocup : {df['cap_gen_ocup'].notna().sum():,}")
print(f"Con cap_gen_tram : {df['cap_gen_tram'].notna().sum():,}")

if args.summary:
    sys.exit(0)

df['fecha'] = df['fecha'].dt.strftime('%Y-%m-%d')
df.to_csv(OUT, index=False, encoding='utf-8-sig')
print(f"Guardado en      : {OUT}")
