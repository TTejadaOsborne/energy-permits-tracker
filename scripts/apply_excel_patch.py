#!/usr/bin/env python3
"""
apply_excel_patch.py
====================
Lee un Excel exportado desde la herramienta con celdas en:
  - ROJO  (FFC00000): datos introducidos manualmente
  - NARANJA (FFFFC000): datos corregidos respecto a la plataforma

Aplica los cambios a los archivos output/energy_extraido_*.json,
actualizando el sub-objeto 'datos{}' de cada proyecto coincidente.
También normaliza gestor_red globalmente en todos los JSONs.

Uso:
  python scripts/apply_excel_patch.py [ruta_excel]
  
  Si no se indica ruta, busca automáticamente el último .xlsx en ~/Downloads
  cuyo nombre contenga 'nodalys' o 'permisos'.
"""

import json
import os
import re
import sys
import glob
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles.fills import PatternFill
except ImportError:
    print("ERROR: openpyxl no instalado. Ejecuta: pip install openpyxl")
    sys.exit(1)


# ── Configuración ───────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent.parent
OUTPUT_DIR = BASE_DIR / "output"

# Colores reconocidos como "dato a aplicar"
RED_COLOR    = "FFC00000"   # rojo — manual
ORANGE_COLOR = "FFFFC000"  # naranja — corregido

# Mapa de columnas Excel → campos en datos{}
COLUMN_MAP = {
    "subestacion":          "subestacion",
    "tension":              "tension",
    "tension_kv":           "tension",
    "gestor_red":           "gestor_red",
    "potencia_mw":          "potencia_mw",
    "potencia":             "potencia_mw",
    "promotor":             "promotor",
    "nombre_proyecto":      "nombre_proyecto",
    "nombre":               "nombre_proyecto",
    "expediente":           "expediente",
    "capacidad_mw_liberada":"capacidad_mw_liberada",
    "capacidad_liberada":   "capacidad_mw_liberada",
}

# Normalización gestor_red
GESTOR_EXACT = {
    "Red Eléctrica de España":                        "REE",
    "Red Eléctrica de España, S.A.U.":                "REE",
    "Red Eléctrica de España, SAU":                   "REE",
    "Red Eléctrica de España (REE)":                  "REE",
    "Red de transporte":                              "REE",
    "REE":                                            "REE",
    "i-DE Redes Eléctricas Inteligentes, S.A.U.":     "Iberdrola",
    "i-DE Redes Eléctricas Inteligentes, SAU":        "Iberdrola",
    "i-DE":                                           "Iberdrola",
    "Iberdrola Distribución Eléctrica, S.A.U.":       "Iberdrola",
    "Iberdrola":                                      "Iberdrola",
    "Endesa Distribución Eléctrica, S.L.":            "Endesa",
    "E-Distribución Redes Digitales, S.L.":           "Endesa",
    "E-Distribución":                                 "Endesa",
    "Endesa":                                         "Endesa",
    "Unión Fenosa Distribución, S.A.":                "UFD/Naturgy",
    "UFD Distribución Electricidad, S.A.":            "UFD/Naturgy",
    "UFD":                                            "UFD/Naturgy",
    "Gas Natural Fenosa":                             "UFD/Naturgy",
    "Viesgo Distribución Eléctrica, S.L.":            "UFD/Naturgy",
    "Viesgo":                                         "UFD/Naturgy",
}

def normalize_gestor(value: str) -> str:
    if not value:
        return value
    v = value.strip()
    if v in GESTOR_EXACT:
        return GESTOR_EXACT[v]
    vl = v.lower()
    if re.search(r'i-?de\s+redes', vl) or re.search(r'iberdrola\s+distribuc', vl):
        return "Iberdrola"
    if re.search(r'red\s+el[eé]ctrica', vl) or 'ree' == vl:
        return "REE"
    if re.search(r'viesgo', vl):
        return "UFD/Naturgy"
    if re.search(r'uni[oó]n\s+fenosa', vl) or re.search(r'\bufd\b', vl):
        return "UFD/Naturgy"
    if re.search(r'endesa\s+distribuc', vl) or re.search(r'e-distribuc', vl):
        return "Endesa"
    return v


# ── Localizar Excel ─────────────────────────────────────────────────────────
def find_excel(arg=None) -> Path:
    if arg:
        p = Path(arg)
        if not p.exists():
            print(f"ERROR: no existe el archivo: {p}")
            sys.exit(1)
        return p

    downloads = Path.home() / "Downloads"
    candidates = sorted(
        [f for f in downloads.glob("*.xlsx")
         if re.search(r'nodalys|permisos', f.name, re.I)],
        key=lambda f: f.stat().st_mtime,
        reverse=True
    )
    if candidates:
        print(f"Excel encontrado automáticamente: {candidates[0]}")
        return candidates[0]

    print("ERROR: No se encontró ningún Excel. Indica la ruta como argumento.")
    sys.exit(1)


# ── Leer patches del Excel ───────────────────────────────────────────────────
def is_colored(cell) -> bool:
    fill = cell.fill
    if not fill or fill.fill_type not in ("solid", "patternFill"):
        return False
    fg = fill.fgColor
    if not fg:
        return False
    color = fg.rgb if fg.type == "rgb" else ""
    return color in (RED_COLOR, ORANGE_COLOR)


def read_patches(excel_path: Path) -> dict:
    """
    Devuelve: {id_boe: {campo: valor, ...}, ...}
    """
    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb.active

    # Leer cabeceras (fila 1)
    headers = []
    for cell in ws[1]:
        val = str(cell.value or "").strip().lower().replace(" ", "_")
        headers.append(val)

    # Buscar índice de id_boe
    id_col = None
    for i, h in enumerate(headers):
        if h in ("id_boe", "id", "boe_id"):
            id_col = i
            break
    if id_col is None:
        print("ERROR: no se encontró columna 'id_boe' en el Excel.")
        sys.exit(1)

    patches = {}
    colored_cells = 0

    for row in ws.iter_rows(min_row=2):
        id_val = row[id_col].value
        if not id_val:
            continue
        id_val = str(id_val).strip()

        for j, cell in enumerate(row):
            if j == id_col or not is_colored(cell):
                continue
            raw_header = headers[j] if j < len(headers) else ""
            field = COLUMN_MAP.get(raw_header)
            if not field:
                continue
            val = cell.value
            if val is None or str(val).strip() == "":
                continue
            val = str(val).strip() if isinstance(val, str) else val

            if id_val not in patches:
                patches[id_val] = {}
            patches[id_val][field] = val
            colored_cells += 1

    print(f"  Celdas coloreadas leídas: {colored_cells}")
    print(f"  Proyectos con patches: {len(patches)}")
    return patches


# ── Aplicar patches a los JSONs ──────────────────────────────────────────────
def apply_patches(patches: dict) -> tuple:
    json_files = sorted(OUTPUT_DIR.glob("energy_extraido_*.json"))
    if not json_files:
        print(f"ERROR: No se encontraron JSONs en {OUTPUT_DIR}")
        sys.exit(1)

    projects_patched = 0
    fields_updated   = 0
    not_found        = set(patches.keys())

    for jf in json_files:
        try:
            with open(jf, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue

        items = data.get("resultados", [])
        changed = False

        for item in items:
            item_id = item.get("id", "")
            if item_id not in patches:
                continue

            not_found.discard(item_id)
            patch = patches[item_id]
            datos = item.setdefault("datos", {})

            for field, value in patch.items():
                old = datos.get(field)
                # Conversión numérica si aplica
                if field in ("potencia_mw", "tension", "capacidad_mw_liberada"):
                    try:
                        value = float(str(value).replace(",", "."))
                    except ValueError:
                        pass
                if old != value:
                    print(f"  [{item_id}] {field}: {repr(old)} → {repr(value)}")
                    datos[field] = value
                    fields_updated += 1
                    changed = True

            projects_patched += 1 if patch else 0

        if changed:
            with open(jf, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

    if not_found:
        print(f"\n⚠  IDs no encontrados en los JSONs ({len(not_found)}):")
        for nf in sorted(not_found):
            print(f"   - {nf}")

    return projects_patched, fields_updated


# ── Normalización global gestor_red ─────────────────────────────────────────
def normalize_all_gestors() -> tuple:
    json_files = sorted(OUTPUT_DIR.glob("energy_extraido_*.json"))
    total_changed = 0
    files_changed = 0

    for jf in json_files:
        try:
            with open(jf, encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            continue

        items = data.get("resultados", [])
        changed = False

        for item in items:
            datos = item.get("datos", {})
            raw = datos.get("gestor_red", "")
            if not raw:
                continue
            norm = normalize_gestor(raw)
            if norm != raw:
                datos["gestor_red"] = norm
                total_changed += 1
                changed = True

        if changed:
            files_changed += 1
            with open(jf, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

    return total_changed, files_changed


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    arg = sys.argv[1] if len(sys.argv) > 1 else None
    excel_path = find_excel(arg)

    print(f"\n{'='*60}")
    print(f"Excel: {excel_path.name}")
    print(f"Output dir: {OUTPUT_DIR}")
    print(f"{'='*60}\n")

    print("1. Leyendo patches del Excel...")
    patches = read_patches(excel_path)

    print("\n2. Aplicando patches a los JSONs...")
    proj_patched, fields_updated = apply_patches(patches)
    print(f"\n   → {proj_patched} proyectos actualizados, {fields_updated} campos modificados")

    print("\n3. Normalizando gestor_red globalmente...")
    g_changed, g_files = normalize_all_gestors()
    print(f"   → {g_changed} valores normalizados en {g_files} archivos")

    print(f"\n{'='*60}")
    print("✓ Patch completo. Próximos pasos:")
    print("  python project_resolver.py")
    print("  python generate_adverse_forecast.py")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
