#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_sets_capacity.py
Genera sets_capacity.json desde Monitor_Capacidad_Red_INTEGRADO_v4.xlsx.

Formato de cada entrada en el JSON:
  key: "NOMBRE kV"  (e.g. "AGUADULC 66", "ABADES 400")
  value: [
    0:  nombre_display       ("AGUADULC 66", "ABADES 400")
    1:  nombre_corto         ("AGUADULC", "ABADES")
    2:  provincia            ("Almería", "" para REE)
    3:  municipio/ccaa       ("Roquetas de Mar", "Castilla y León")
    4:  tension_kv           (66.0, 400.0)
    5:  cap_actual_mw        (puede ser null)
    6:  cap_anterior_mw      (puede ser null)
    7:  cap_ocupada_mw       (puede ser null)
    8:  acept_actual         (puede ser null)
    9:  acept_anterior       (puede ser null)
    10: titular              ("Endesa", "iDE", "UFD", "Viesgo", "REE", ...)
    11: red                  ("Distribución" o "Transporte")
    12: afloramiento         (cap_actual - cap_anterior, puede ser null)
    13: lat                  (WGS84, null si no hay coords)
    14: lon                  (WGS84, null si no hay coords)
    15: cap_demanda_actual   (MW demanda disponible — DSO col21, REE col7)
    16: afloramiento_demanda (cap_demanda_actual - cap_demanda_anterior)
  ]

Uso:
  python generate_sets_capacity.py
  python generate_sets_capacity.py --excel Monitor_Capacidad_Red_INTEGRADO_v4.xlsx --out sets_capacity.json
  python generate_sets_capacity.py --kml-ree "Red de Transporte.txt"   # incluye coords REE del KML
"""

import argparse
import json
import os
import re
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: Instala openpyxl:  pip install openpyxl")
    sys.exit(1)

try:
    from pyproj import Transformer
    _T30 = Transformer.from_crs('EPSG:25830', 'EPSG:4326', always_xy=True)  # Zone 30N (España peninsular + Baleares)
    _T28 = Transformer.from_crs('EPSG:25828', 'EPSG:4326', always_xy=True)  # Zone 28N (Canarias)
    _HAS_PYPROJ = True
except ImportError:
    _HAS_PYPROJ = False
    print("AVISO: pyproj no disponible — lat/lon no se calcularán. pip install pyproj")

PROVINCIAS_CANARIAS = {'palmas, las', 'santa cruz de tenerife'}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def to_float(v):
    """Convierte valor de celda a float, None si no aplica."""
    if v is None or v == "" or v == "N/A":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def calc_afloramiento(actual, anterior):
    """Afloramiento = actual - anterior. None si falta algún dato."""
    if actual is None or anterior is None:
        return None
    return round(actual - anterior, 4)


def utm_to_latlon(x, y, provincia: str):
    """Convierte UTM ETRS89 a WGS84. Usa Zone 28 para Canarias, 30 para el resto."""
    if not _HAS_PYPROJ or x is None or y is None:
        return None, None
    try:
        transformer = _T28 if (provincia or "").lower() in PROVINCIAS_CANARIAS else _T30
        lon, lat = transformer.transform(float(x), float(y))
        # Sanity check: España continental + Canarias
        if not (-18 <= lon <= 5 and 27 <= lat <= 44):
            return None, None
        return round(lat, 6), round(lon, 6)
    except Exception:
        return None, None


def extract_tension(nombre_nudo: str):
    """Extrae kV del nombre del nudo REE: 'ABADES 400' → 400.0"""
    tokens = nombre_nudo.strip().split()
    if not tokens:
        return None
    last = tokens[-1]
    try:
        return float(last)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Lectura KML REE (coordenadas)
# ---------------------------------------------------------------------------

def parse_kml_ree_coords(kml_path):
    """
    Lee un KML de REE y devuelve {nombre_nudo: (lat, lon)}.
    Hace matching con 4 estrategias para cubrir nombres con paréntesis/alias.
    """
    def _norm(s):
        s = re.sub(r'\s*\([^)]*\)', '', s)
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    tree = ET.parse(kml_path)
    root = tree.getroot()
    placemarks = {}
    for pm in root.iter('{http://www.opengis.net/kml/2.2}Placemark'):
        name_el   = pm.find('{http://www.opengis.net/kml/2.2}name')
        coords_el = pm.find('.//{http://www.opengis.net/kml/2.2}coordinates')
        if name_el is not None and coords_el is not None:
            name = name_el.text.strip()
            parts = coords_el.text.strip().split(',')
            if len(parts) >= 2:
                placemarks[name] = (round(float(parts[1]), 6), round(float(parts[0]), 6))

    print(f"  KML REE: {len(placemarks)} placemarks leídos")
    return placemarks


def build_ree_coord_lookup(placemarks, ree_keys):
    """Construye {ree_key: (lat, lon)} usando 4 estrategias de matching."""
    norm_kml = {}
    for name, coords in placemarks.items():
        n = _norm_key(name)
        if n not in norm_kml:
            norm_kml[n] = coords

    coord_map = {}
    for k in ree_keys:
        if k in placemarks:
            coord_map[k] = placemarks[k]; continue
        nk = _norm_key(k)
        if nk in norm_kml:
            coord_map[k] = norm_kml[nk]; continue
        m = re.match(r'^(.+?)\s*\((.+?)\)\s*(\d+)$', k)
        if m:
            alias_key = m.group(2).strip() + ' ' + m.group(3).strip()
            if alias_key in placemarks:
                coord_map[k] = placemarks[alias_key]; continue
            if _norm_key(alias_key) in norm_kml:
                coord_map[k] = norm_kml[_norm_key(alias_key)]; continue
        if k.endswith(' (REE)'):
            base = k[:-6]
            if base in placemarks:
                coord_map[k] = placemarks[base]; continue
            if _norm_key(base) in norm_kml:
                coord_map[k] = norm_kml[_norm_key(base)]; continue

    print(f"  KML REE matching: {len(coord_map)}/{len(ree_keys)} nudos con coords")
    return coord_map


def _norm_key(s):
    """Normaliza clave para matching: quita paréntesis, espacios extra."""
    s = re.sub(r'\s*\([^)]*\)', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# ---------------------------------------------------------------------------
# Lectura DSO Actual
# ---------------------------------------------------------------------------

def read_dso_actual(ws):
    """
    DSO Actual: headers en fila 5, datos desde fila 6.
    Columnas relevantes (0-indexadas):
      1  Nombre DSO (titular)
      3  Provincia
      4  Municipio
      5  Nombre Subestación (nombre corto)
      9  Tensión (kV)
      10 [SET + kV]  ← clave
      11 Cap firme disponible actual (MW)
      12 Cap firme disponible previa (MW)
      15 Cap de acceso firme de demanda ocupada (MW)
      34 Aceptabilidad Neta REE actual
      35 Aceptabilidad Neta REE previa
    """
    entries = {}
    n_skipped = 0

    for row in ws.iter_rows(min_row=6, values_only=True):
        key = row[10]  # [SET + kV]
        if not key or not str(key).strip():
            n_skipped += 1
            continue

        key = str(key).strip()

        titular_raw  = str(row[1] or "").strip()
        provincia    = str(row[3] or "").strip()
        municipio    = str(row[4] or "").strip()
        nombre_set   = str(row[5] or "").strip()
        tension      = to_float(row[9])

        cap_actual    = to_float(row[11])
        cap_anterior  = to_float(row[12])
        cap_ocupada   = to_float(row[15])
        cap_dem_actual  = to_float(row[21])
        cap_dem_anterior = to_float(row[22])
        acept_actual  = to_float(row[34])
        acept_anterior = to_float(row[35])

        afloramiento     = calc_afloramiento(cap_actual, cap_anterior)
        aflor_demanda    = calc_afloramiento(cap_dem_actual, cap_dem_anterior)
        titular      = normalizar_titular(titular_raw)
        lat, lon     = utm_to_latlon(row[7], row[8], provincia)

        nombre_display = key
        nombre_corto   = nombre_set or key.rsplit(" ", 1)[0]

        entries[key] = [
            nombre_display,  # 0
            nombre_corto,    # 1
            provincia,       # 2
            municipio,       # 3
            tension,         # 4
            cap_actual,      # 5
            cap_anterior,    # 6
            cap_ocupada,     # 7
            acept_actual,    # 8
            acept_anterior,  # 9
            titular,         # 10
            "Distribución",  # 11
            afloramiento,    # 12
            lat,             # 13
            lon,             # 14
            cap_dem_actual,  # 15
            aflor_demanda,   # 16
        ]

    print(f"  DSO Actual: {len(entries)} entradas leídas ({n_skipped} filas vacías omitidas)")
    return entries


# ---------------------------------------------------------------------------
# Lectura REE Actual
# ---------------------------------------------------------------------------

def read_ree_actual(ws, ree_coord_map=None):
    """
    REE Actual: datos desde fila 9.
    Columnas relevantes (0-indexadas):
      1  Nombre y tensión del nudo  ← clave ("ABADES 400")
      3  Comunidad Autónoma
      4  NETO GENERACIÓN ALM (actual)
      5  NETO GENERACIÓN PREVIO
      10 ACEPTABILIDAD NETA
    """
    entries = {}
    n_skipped = 0

    for row in ws.iter_rows(min_row=9, values_only=True):
        nombre_nudo = row[1]
        if not nombre_nudo or not str(nombre_nudo).strip():
            n_skipped += 1
            continue

        nombre_nudo  = str(nombre_nudo).strip()
        key          = nombre_nudo
        ccaa         = str(row[3] or "").strip()
        cap_actual   = to_float(row[4])
        cap_anterior = to_float(row[5])
        cap_dem_actual  = to_float(row[7])
        cap_dem_anterior = to_float(row[8])
        acept_actual = to_float(row[10])

        afloramiento  = calc_afloramiento(cap_actual, cap_anterior)
        aflor_demanda = calc_afloramiento(cap_dem_actual, cap_dem_anterior)
        tension      = extract_tension(nombre_nudo)

        tokens       = nombre_nudo.split()
        nombre_corto = " ".join(tokens[:-1]) if len(tokens) > 1 else nombre_nudo

        entries[key] = [
            nombre_nudo,    # 0  nombre_display
            nombre_corto,   # 1  nombre_corto
            "",             # 2  provincia (no disponible en REE)
            ccaa,           # 3  ccaa como municipio
            tension,        # 4  tension_kv
            cap_actual,     # 5  cap_actual
            cap_anterior,   # 6  cap_anterior
            None,           # 7  cap_ocupada (no aplica REE)
            acept_actual,   # 8  acept_actual
            acept_actual,   # 9  acept_anterior (no hay columna previa para aceptabilidad en REE)
            "REE",          # 10 titular
            "Transporte",   # 11 red
            afloramiento,   # 12 afloramiento
            ree_coord_map.get(key, (None,None))[0] if ree_coord_map else None,  # 13 lat
            ree_coord_map.get(key, (None,None))[1] if ree_coord_map else None,  # 14 lon
            cap_dem_actual,  # 15
            aflor_demanda,   # 16
        ]

    print(f"  REE Actual: {len(entries)} entradas leídas ({n_skipped} filas vacías omitidas)")
    return entries


# ---------------------------------------------------------------------------
# Resolución de colisiones DSO vs REE
# ---------------------------------------------------------------------------

def merge_and_resolve_collisions(dso_entries, ree_entries):
    """
    Combina DSO + REE. Cuando una clave existe en ambos,
    la entrada REE recibe sufijo ' (REE)'.
    """
    result = dict(dso_entries)
    collisions = 0

    for key, value in ree_entries.items():
        if key in result:
            new_key = key + " (REE)"
            value = list(value)
            value[0] = new_key
            result[new_key] = value
            collisions += 1
            print(f"  Colisión: '{key}' → REE como '{new_key}'")
        else:
            result[key] = value

    print(f"  {collisions} colisiones resueltas.")
    return result


# ---------------------------------------------------------------------------
# Normalización de titular
# ---------------------------------------------------------------------------

TITULAR_MAP = [
    ("i-de",           "iDE"),
    ("i de ",          "iDE"),
    ("iberdrola distribución", "iDE"),
    ("iberdrola distribucion", "iDE"),
    ("ide",            "iDE"),   # covers "iDE", "IDE", "ide"
    ("endesa",         "Endesa"),
    ("e-distribución", "Endesa"),
    ("edistribucion",  "Endesa"),
    ("e distribución", "Endesa"),
    ("ufd",            "UFD"),
    ("unión fenosa",   "UFD"),
    ("union fenosa",   "UFD"),
    ("viesgo",         "Viesgo"),
    ("red eléctrica",  "REE"),
    ("red electrica",  "REE"),
    ("ree",            "REE"),
    ("naturgy",        "Naturgy"),
]

def normalizar_titular(raw: str) -> str:
    if not raw:
        return raw
    lower = raw.lower()
    for pattern, canonical in TITULAR_MAP:
        if pattern in lower:
            return canonical
    return raw.title()


# ---------------------------------------------------------------------------
# Stats
# ---------------------------------------------------------------------------

def print_stats(data):
    total     = len(data)
    con_cap   = sum(1 for v in data.values() if v[5] is not None)
    con_aflor = sum(1 for v in data.values() if v[12] is not None)
    con_coords = sum(1 for v in data.values() if len(v) > 13 and v[13] is not None)
    aflor_pos = sum(1 for v in data.values() if v[12] and v[12] > 0)
    aflor_neg = sum(1 for v in data.values() if v[12] and v[12] < 0)

    titulares = {}
    for v in data.values():
        t = v[10]
        titulares[t] = titulares.get(t, 0) + 1

    print(f"\n  Total entradas:       {total}")
    print(f"  Con cap_actual:       {con_cap} ({100*con_cap//total}%)")
    print(f"  Con afloramiento:     {con_aflor} ({100*con_aflor//total}%)")
    print(f"  Afloramiento > 0:     {aflor_pos}")
    ree_with = sum(1 for v in data.values() if v[10]=='REE' and len(v)>13 and v[13] is not None)
    print(f"  Con lat/lon:          {con_coords} (REE: {ree_with}/937)")
    print(f"  Afloramiento < 0:     {aflor_neg}")
    print(f"  Por titular:")
    for t, n in sorted(titulares.items(), key=lambda x: -x[1]):
        print(f"    {t:25s}: {n}")

    top = sorted(
        [(k, v[12]) for k, v in data.items() if v[12] and v[12] > 0],
        key=lambda x: -x[1]
    )[:10]
    if top:
        print(f"\n  Top 10 afloramientos positivos (MW):")
        for k, a in top:
            print(f"    {k:35s}: +{a:.1f} MW")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Genera sets_capacity.json desde el Excel de monitorización."
    )
    parser.add_argument(
        "--excel", default="Monitor_Capacidad_Red_INTEGRADO_v4.xlsx",
        help="Ruta al Excel (default: Monitor_Capacidad_Red_INTEGRADO_v4.xlsx)"
    )
    parser.add_argument(
        "--out", default="sets_capacity.json",
        help="Archivo de salida (default: sets_capacity.json)"
    )
    parser.add_argument(
        "--kml-ree", default=None,
        help="KML con coordenadas de subestaciones REE (Red de Transporte.txt)"
    )
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"ERROR: No se encuentra el Excel: {excel_path}")
        sys.exit(1)

    print(f"Leyendo {excel_path} ...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    for sheet in ("DSO Actual", "REE Actual"):
        if sheet not in wb.sheetnames:
            print(f"ERROR: No se encuentra la hoja '{sheet}'")
            sys.exit(1)

    print("\nLeyendo DSO Actual...")
    dso_entries = read_dso_actual(wb["DSO Actual"])

    ree_coord_map = None
    if args.kml_ree:
        kml_path = Path(args.kml_ree)
        if kml_path.exists():
            print("\nLeyendo KML REE...")
            kml_placemarks = parse_kml_ree_coords(kml_path)
            ree_coord_map_raw = {}  # temp — needed after REE keys are known
        else:
            print(f"AVISO: KML no encontrado: {kml_path}")

    print("\nLeyendo REE Actual...")
    ree_entries_raw = read_ree_actual(wb["REE Actual"])

    # Ahora que tenemos las claves REE, construir lookup de coords
    if args.kml_ree and ree_coord_map is None and 'kml_placemarks' in dir():
        ree_coord_map = build_ree_coord_lookup(kml_placemarks, set(ree_entries_raw.keys()))
        # Re-leer con coords
        ree_entries = read_ree_actual(wb["REE Actual"], ree_coord_map)
    else:
        ree_entries = ree_entries_raw

    print("\nMerging y resolviendo colisiones...")
    data = merge_and_resolve_collisions(dso_entries, ree_entries)

    print("\nEstadísticas:")
    print_stats(data)

    out_path = Path(args.out)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

    size_kb = out_path.stat().st_size / 1024
    print(f"\n✓ Guardado: {out_path}  ({size_kb:.0f} KB, {len(data)} entradas)")


if __name__ == "__main__":
    main()
