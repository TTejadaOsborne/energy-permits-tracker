#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""generate_sets_history.py v3 — Monitor Excel + CSVs históricos DSO individuales"""

import argparse, json, re, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl"); sys.exit(1)
try:
    import pandas as pd
except ImportError:
    print("pip install pandas"); sys.exit(1)

MESES_ES = {"ene":1,"feb":2,"mar":3,"abr":4,"may":5,"jun":6,
            "jul":7,"ago":8,"sep":9,"oct":10,"nov":11,"dic":12}
HIST_RE  = re.compile(r'^(DSO|REE)\s+([A-Za-z]+)(\d{2})$', re.IGNORECASE)

# CSVs de parsers individuales (solo generación; demanda = null)
DSO_CSV_SOURCES = {
    "endesa_capacidad.csv":  "Endesa",
    "eredes_capacidad.csv":  "ERedes",
    "ide_capacidad.csv":     "iDE",
    "ree_capacidad.csv":     "REE",
    "ufd_capacidad.csv":     "UFD",
    "viesgo_capacidad.csv":  "Viesgo",
}

def to_float(v):
    if v is None or v == "" or v == "N/A": return None
    try:
        f = float(v)
    except:
        return None
    return None if (f != f) else f  # NaN check (f != f is True only for NaN)

def add_floats(*vals):
    nums = [v for v in vals if v is not None]
    return round(sum(nums), 4) if nums else None

def parse_sheet(name):
    m = HIST_RE.match(name.strip())
    if not m: return None
    tipo = m.group(1).upper()
    mes  = MESES_ES.get(m.group(2).lower()[:3])
    year = 2000 + int(m.group(3))
    return (tipo, year, mes) if mes else None

def label(y, m): return f"{y:04d}-{m:02d}"

def extract_dso(ws, year, mes):
    entries = {}
    lbl = label(year, mes)

    # Buscar índice de columna "Capacidad de acceso disponible para MPE RdD" y "MPE RdD" en headers (filas 4-5)
    rdd_col_idx = None
    acept_col_idx = None
    for header_row_num in [5, 4]:  # Buscar en fila 5 primero, luego fila 4
        header_row = list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))[0] if ws.max_row >= header_row_num else None
        if header_row:
            for i, cell in enumerate(header_row):
                if cell and isinstance(cell, str):
                    if "Capacidad de acceso disponible para MPE RdD" in cell and rdd_col_idx is None:
                        rdd_col_idx = i
                    if "MPE RdD" in cell and "MGES" not in cell and acept_col_idx is None:
                        acept_col_idx = i
            if rdd_col_idx or acept_col_idx:
                break

    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row or len(row) < 27: continue
        key = row[10]
        if not key or not str(key).strip(): continue
        key = str(key).strip()
        _cg = to_float(row[21]); _cd = to_float(row[11])
        # cap_gen_RdD: usar índice encontrado en header
        _cap_gen_RdD = None
        if rdd_col_idx and len(row) > rdd_col_idx:
            _cap_gen_RdD = to_float(row[rdd_col_idx])

        # acept = Cap. acceso disponible MPE RdD - Cap. solicitada en curso MPE
        _tram = to_float(row[26])
        _acept = None
        if _cap_gen_RdD is not None and _tram is not None:
            _acept = round(_cap_gen_RdD - _tram, 4)

        entries[key] = {
            "date":         lbl,
            "cap_gen":      _cg,
            "cap_dem":      _cd,
            "cap_gen_ocup": to_float(row[25]),
            "cap_gen_tram": to_float(row[26]),
            "cap_dem_ocup": to_float(row[15]),
            "cap_dem_tram": to_float(row[16]),
            "cap_gen_disp": _cg,
            "cap_gen_neta": _cg,
            "cap_dem_disp": _cd,
            "cap_dem_neta": _cd,
            "acept":        _acept,
            "cap_gen_RdD":  _cap_gen_RdD,
        }
    return entries

def extract_ree(ws, year, mes):
    entries = {}
    lbl = label(year, mes)

    # Buscar col "MPE RdD" (grupo Generacion = primera ocurrencia) en filas 1-8
    rdd_col = None
    for hrow in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        for i, h in enumerate(hrow):
            hh = " ".join(str(h or "").lower().split())
            if "mpe rdd" in hh and "mges" not in hh and "no conectado" not in hh:
                rdd_col = i
                break
        if rdd_col is not None:
            break

    for row in ws.iter_rows(min_row=9, values_only=True):
        if not row or len(row) < 102: continue
        key = row[1]
        if not key or not str(key).strip(): continue
        key = str(key).strip()
        _gen_disp = to_float(row[44]) if len(row) > 44 else None
        _gen_tram = to_float(row[40])
        _rdd = to_float(row[rdd_col]) if (rdd_col is not None and len(row) > rdd_col) else None
        _cd = to_float(row[7])
        entries[key] = {
            "date":         lbl,
            "cap_gen":      to_float(row[4]),
            "cap_dem":      _cd,
            "cap_gen_ocup": to_float(row[33]),
            "cap_gen_tram": _gen_tram,
            "cap_dem_ocup": add_floats(to_float(row[89]), to_float(row[93])),
            "cap_dem_tram": add_floats(to_float(row[100]), to_float(row[101])),
            "cap_gen_disp": _gen_disp,
            "cap_gen_neta": (round(_gen_disp - (_gen_tram or 0.0), 4) if _gen_disp is not None else None),
            "cap_dem_disp": None,
            "cap_dem_neta": _cd,
            # acept: columna precalculada "ACEPTABILIDAD NETA" del Monitor
            "acept":        to_float(row[10]),
            "cap_gen_RdD":  _rdd,
        }
    return entries

def extract_from_mapas_ree(wb_mapas, sheet_name, year, mes_num):
    """Busca automaticamente columnas RdD y Tram por header.

    Soporta los 3 formatos historicos del archivo Mapas:
    - 2022 temprano: headers en fila 5, sin desglose RdT/RdD
      (fallback: "Capacidad de acceso disponible para MPE")
    - oct 2022 - 2024: headers en fila 4, col "...disponible para MPE RdD"
    - 2025+: headers en fila 4, "MPE RdD" duplicada en grupos
      Generacion/Almacenamiento (se toma la primera = Generacion)
    """
    date_key = f"{year:04d}-{mes_num:02d}"
    entries = {}

    if sheet_name not in wb_mapas.sheetnames:
        return entries

    ws = wb_mapas[sheet_name]

    # 1) Escanear filas 1-8: headers pueden estar repartidos en varias filas
    #    (ej. '20 junio 2022': nombre en fila 2, columnas RdD/Tram en fila 3)
    scanned = []
    for ri, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        scanned.append((ri, [" ".join(str(v or "").lower().split()) for v in row]))

    name_col = None
    name_row = None
    for ri, vals in scanned:
        for i_, h in enumerate(vals):
            if "nombre y tensi" in h:
                name_col, name_row = i_, ri
                break
        if name_col is not None:
            break
    if name_col is None:
        return entries

    # 2) Columna RdD: preferir "mpe rdd" exclusivo de generacion
    #    (la primera ocurrencia pertenece al grupo Generacion en formato 2025)
    rdd_col = rdd_row = None
    for ri, vals in scanned:
        for i_, h in enumerate(vals):
            if "mpe rdd" in h and "mges" not in h and "no conectado" not in h:
                rdd_col, rdd_row = i_, ri
                break
        if rdd_col is not None:
            break
    if rdd_col is None:
        # Fallback formato 2022 temprano: sin desglose RdT/RdD
        for ri, vals in scanned:
            for i_, h in enumerate(vals):
                if "capacidad de acceso disponible para mpe" in h and "rdt" not in h:
                    rdd_col, rdd_row = i_, ri
                    break
            if rdd_col is not None:
                break

    # 3) Columna Tram: "...solicitada en curso y pendiente resolver MPE" (exacta)
    tram_col = tram_row = None
    for ri, vals in scanned:
        for i_, h in enumerate(vals):
            hh = h.replace("[mw]", "").strip()
            if "capacidad de acceso solicitada" in hh and hh.endswith("resolver mpe"):
                tram_col, tram_row = i_, ri
                break
        if tram_col is not None:
            break

    if rdd_col is None or tram_col is None:
        return entries

    data_start = max(name_row, rdd_row, tram_row) + 1
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or len(row) < max(rdd_col, tram_col, name_col) + 1:
            continue

        key = row[name_col] if row[name_col] else None
        if not key or not str(key).strip():
            continue
        key = str(key).strip()

        cap_gen_rdd = to_float(row[rdd_col])
        cap_gen_tram = to_float(row[tram_col])

        acept = None
        if cap_gen_rdd is not None and cap_gen_tram is not None:
            acept = round(cap_gen_rdd - cap_gen_tram, 4)

        entries[key] = {
            "date": date_key,
            "cap_gen": None,
            "cap_dem": None,
            "cap_gen_ocup": None,
            "cap_gen_tram": cap_gen_tram,
            "cap_dem_ocup": None,
            "cap_dem_tram": None,
            "cap_gen_disp": None,
            "cap_gen_neta": None,
            "cap_dem_disp": None,
            "cap_dem_neta": None,
            "acept": acept,
            "cap_gen_RdD": cap_gen_rdd,
        }

    return entries

def parse_special_sheet(wb, name):
    """Hojas 'REE Actual'/'REE Anterior'/'DSO Actual'/'DSO Anterior': el mes/año
    se leen de la celda B3 (datetime)."""
    if name not in ("REE Actual","REE Anterior","DSO Actual","DSO Anterior"):
        return None
    tipo = "REE" if name.startswith("REE") else "DSO"
    b3 = wb[name]["B3"].value
    if b3 is None: return None
    try:
        year, mes = b3.year, b3.month
    except AttributeError:
        return None
    return (tipo, year, mes)

def build_from_monitor(wb, raw):
    """Lee hojas DSO Xxx / REE Xxx del Monitor y puebla raw."""
    hist = [(y,m,t,n) for n in wb.sheetnames
            for res in [parse_sheet(n)] if res
            for t,y,m in [res]]
    # Hojas "Actual"/"Anterior": mes/año vienen de B3
    for n in wb.sheetnames:
        res = parse_special_sheet(wb, n)
        if res:
            t,y,m = res
            hist.append((y,m,t,n))
    if not hist:
        print("  Sin hojas históricas en el Monitor."); return
    hist.sort(key=lambda x:(x[0],x[1]))
    print(f"\nMonitor: {len(hist)} hojas históricas")
    for y,m,t,n in hist:
        ws = wb[n]
        lbl = label(y,m)
        entries = extract_dso(ws,y,m) if t=="DSO" else extract_ree(ws,y,m)
        print(f"  '{n}': {len(entries)} entradas")
        for key, snap in entries.items():
            if key not in raw: raw[key] = {}
            if lbl in raw[key]:
                ex = raw[key][lbl]
                for f in ("cap_gen","cap_dem","cap_gen_ocup","cap_gen_tram",
                          "cap_dem_ocup","cap_dem_tram","cap_gen_disp","cap_gen_neta",
                          "cap_dem_disp","cap_dem_neta","acept"):
                    if ex[f] is None and snap[f] is not None:
                        ex[f] = snap[f]
            else:
                raw[key][lbl] = snap

def build_from_csvs(refs_dir, raw):
    """
    Lee CSVs individuales de cada DSO (solo generación).
    Columna clave: 'subestacion' + 'tension_kv'  -> key = "NOMBRE kV"
    Solo añade snapshots que no existen ya en raw (el Monitor tiene prioridad).
    Demanda siempre null (estos archivos no la contienen).
    """
    total_added = 0
    for fname, dist in DSO_CSV_SOURCES.items():
        path = refs_dir / fname
        if not path.exists():
            print(f"  SKIP {dist}: {fname} no encontrado")
            continue
        df = pd.read_csv(path, encoding="utf-8-sig", dtype=str)
        if "subestacion" not in df.columns or "disponible_neta" not in df.columns:
            print(f"  SKIP {dist}: columnas inesperadas en {fname}")
            continue

        df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
        df = df.dropna(subset=["fecha"])
        df["tension_kv"] = pd.to_numeric(df["tension_kv"], errors="coerce")

        has_ocup = "cap_gen_ocup" in df.columns
        has_tram = "cap_gen_tram" in df.columns
        has_disp = "disponible_bruta" in df.columns

        added = 0
        for row in df.itertuples(index=False):
            r = row._asdict()
            nombre = str(r["subestacion"]).strip().upper()
            if not nombre or nombre == "NAN": continue
            kv = r["tension_kv"]
            kv_str = f"{int(kv)}" if pd.notna(kv) and kv == int(kv) else (f"{kv}" if pd.notna(kv) else "")
            key = f"{nombre} {kv_str}".strip()
            fecha = r["fecha"]
            lbl = label(fecha.year, fecha.month)

            cap_gen = to_float(r["disponible_neta"])
            # Ocupada/trámite si el CSV las tiene (ej: ree_capacidad.csv)
            cap_gen_ocup = to_float(r["cap_gen_ocup"]) if has_ocup else None
            cap_gen_tram = to_float(r["cap_gen_tram"]) if has_tram else None
            cap_gen_disp = to_float(r["disponible_bruta"]) if has_disp else None
            _cap_gen_RdD = cap_gen  # En CSVs, disponible_neta es cap_gen_RdD
            _acept = (cap_gen - cap_gen_tram) if (cap_gen is not None and cap_gen_tram is not None) else None
            snap = {
                "date":         lbl,
                "cap_gen":      cap_gen,
                "cap_dem":      None,   # no disponible en Excels DSO individuales
                "cap_gen_ocup": cap_gen_ocup,
                "cap_gen_tram": cap_gen_tram,
                "cap_dem_ocup": None,
                "cap_dem_tram": None,
                "cap_gen_disp": cap_gen_disp,
                "cap_gen_neta": cap_gen,
                "cap_dem_disp": None,
                "cap_dem_neta": None,
                "acept":        _acept,
                "cap_gen_RdD":  _cap_gen_RdD,
            }
            if key not in raw: raw[key] = {}
            if lbl in raw[key]:
                # Ya existe (Monitor o Mapas): rellenar solo campos null,
                # sin sobreescribir (Monitor/Mapas tienen prioridad)
                ex = raw[key][lbl]
                merged = False
                for f, v in snap.items():
                    if f != "date" and ex.get(f) is None and v is not None:
                        ex[f] = v
                        merged = True
                if merged:
                    added += 1
            else:
                raw[key][lbl] = snap
                added += 1

        total_added += added
        print(f"  {dist} ({fname}): +{added} snapshots añadidos (gen only, dem=null)")

    print(f"  Total desde CSVs individuales: +{total_added} snapshots")



def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel",   default="Monitor_Capacidad_Red_INTEGRADO_v4.xlsx")
    parser.add_argument("--refs",    default="references")
    parser.add_argument("--out",     default="sets_history.json")
    parser.add_argument("--no-csvs", action="store_true", help="Solo Monitor, sin CSVs individuales")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    refs_dir   = Path(args.refs)

    raw = {}  # key -> {date_label: snap}

    # 1) Monitor Excel (generación + demanda, Oct-2025 en adelante)
    if excel_path.exists():
        print(f"Leyendo Monitor: {excel_path} ...")
        wb = openpyxl.load_workbook(excel_path, read_only=False, data_only=True)
        build_from_monitor(wb, raw)
        wb.close()
    else:
        print(f"WARN: Monitor no encontrado en {excel_path}")

    mapas_path = Path("references") / "Mapas_Capacidad_AyC-REE.xlsx"
    if mapas_path.exists():
        print(f"Leyendo históricos REE: {mapas_path} ...")
        wb_mapas = openpyxl.load_workbook(mapas_path, read_only=False, data_only=True)
        for sheet_name in wb_mapas.sheetnames:
            try:
                sheet_lower = sheet_name.lower().strip()
                meses = {"enero":1,"febrero":2,"marzo":3,"abril":4,"mayo":5,"junio":6,
                        "julio":7,"agosto":8,"septiembre":9,"octubre":10,"noviembre":11,"diciembre":12}
                for mes_name, mes_num in meses.items():
                    if mes_name in sheet_lower:
                        year_match = re.search(r'(202[0-9]|201[0-9])', sheet_name)
                        if year_match:
                            year = int(year_match.group(1))
                            entries = extract_from_mapas_ree(wb_mapas, sheet_name, year, mes_num)
                            for key, snap in entries.items():
                                if key not in raw: raw[key] = {}
                                if snap["date"] in raw[key]:
                                    # Merge: actualizar solo campos null
                                    for f in ("cap_gen_tram","acept","cap_gen_RdD"):
                                        if raw[key][snap["date"]][f] is None and snap[f] is not None:
                                            raw[key][snap["date"]][f] = snap[f]
                                else:
                                    raw[key][snap["date"]] = snap
                        break
            except:
                pass
        wb_mapas.close()
    else:
        print(f"WARN: Mapas_Capacidad_AyC-REE.xlsx no encontrado")

    # 2) CSVs individuales DSO (solo generación, dem=null, 2021-2025)
    if not args.no_csvs and refs_dir.exists():
        print(f"\nLeyendo CSVs históricos en {refs_dir} ...")
        build_from_csvs(refs_dir, raw)

    if not raw:
        print("ERROR: sin datos."); sys.exit(1)

    history_objs = {k: sorted(v.values(), key=lambda s: s["date"]) for k, v in raw.items()}

    # Stats
    print(f"\n--- RESUMEN ---")
    print(f"SETs únicos      : {len(history_objs):,}")
    total_snaps = sum(len(v) for v in history_objs.values())
    print(f"Snapshots totales: {total_snaps:,}")
    all_dates = sorted({s["date"] for v in history_objs.values() for s in v})
    print(f"Rango fechas     : {all_dates[0]} -> {all_dates[-1]}")
    has_dem = sum(1 for v in history_objs.values() if any(s["cap_dem"] is not None for s in v))
    has_gen = sum(1 for v in history_objs.values() if any(s["cap_gen"] is not None for s in v))
    print(f"SETs con cap_dem : {has_dem:,}")
    print(f"SETs con cap_gen : {has_gen:,}")

    # Formato compacto: arrays en vez de objetos
    # [0]=date [1]=cap_gen [2]=cap_gen_ocup [3]=cap_gen_tram
    # [4]=cap_dem [5]=cap_dem_ocup [6]=cap_dem_tram [7]=acept
    # [8]=cap_gen_disp [9]=cap_gen_neta [10]=cap_dem_disp [11]=cap_dem_neta
    # [12]=cap_gen_RdD
    _FIELDS = ("date","cap_gen","cap_gen_ocup","cap_gen_tram",
               "cap_dem","cap_dem_ocup","cap_dem_tram","acept",
               "cap_gen_disp","cap_gen_neta","cap_dem_disp","cap_dem_neta","cap_gen_RdD")
    history = {"_v": 2}
    for k, snaps in history_objs.items():
        history[k] = [[s[f] for f in _FIELDS] for s in snaps]
    out = Path(args.out)
    try:
        import orjson
        with open(out, "wb") as f:
            f.write(orjson.dumps(history))
    except ImportError:
        with open(out, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, separators=(",", ":"))
    print(f"\nOK {out}  ({out.stat().st_size/1024:.0f} KB, {len(history_objs)} SETs)")

if __name__ == "__main__":
    main()
