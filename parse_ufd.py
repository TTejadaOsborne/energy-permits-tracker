"""
parse_ufd.py
Extrae la capacidad disponible neta para generación de UFD desde
el Excel 'references/Mapas_Capacidad_AyC-UFD.xlsx'.

Lógica:
  - Formato estándar (todas las hojas salvo Oct-2025):
      * Una fila por posición de parque; se toma la fila 'Total parque' por SET
      * Si disponible vacío → max del grupo
      * disponible_neta = disponible - (FV_estudio + eolico_estudio + otros_estudio)
  - Formato Oct-2025:
      * Una fila por SET, sin desglose FV/Eólico/Otros
      * disponible_neta = capacidad_disponible - admitida_no_resuelta

Salida: references/ufd_capacidad.csv
  fecha, distribuidora, provincia, municipio, subestacion, tension_kv, disponible_neta

Uso:
  python parse_ufd.py
  python parse_ufd.py --summary
"""
import re
import sys
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("pip install openpyxl --break-system-packages")
try:
    import pandas as pd
except ImportError:
    sys.exit("pip install pandas --break-system-packages")

BASE = Path(__file__).parent
XLSX = BASE / "references" / "Mapas_Capacidad_AyC-UFD.xlsx"
OUT  = BASE / "references" / "ufd_capacidad.csv"

MESES = {
    'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
    'julio':7,'agosto':8,'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12
}
MESES_ABBR = {
    'ene':'enero','feb':'febrero','mar':'marzo','abr':'abril','may':'mayo','jun':'junio',
    'jul':'julio','ago':'agosto','sep':'septiembre','oct':'octubre','nov':'noviembre','dic':'diciembre',
}


def parse_sheet_date(name):
    name = name.strip()
    m = re.match(r'(\d{1,2})\s+(\w+)\s+(\d{4})', name)
    if m:
        day, mes, year = int(m.group(1)), m.group(2).lower(), int(m.group(3))
        if mes in MESES:
            return f'{year}-{MESES[mes]:02d}-{day:02d}'
    m2 = re.match(r'([A-Za-z]{3,})[- _](\d{4})', name)
    if m2:
        raw, year = m2.group(1).lower(), int(m2.group(2))
        mes_full = MESES_ABBR.get(raw[:3], raw)
        if mes_full in MESES:
            return f'{year}-{MESES[mes_full]:02d}-01'
    return None


def to_float(v):
    if v is None or str(v).strip() in ('', 'N/A', 'N/D', '-', 'NA', 'Not found'):
        return None
    try:
        return float(str(v).replace(',', '.'))
    except (ValueError, TypeError):
        return None


def detect_format_and_cols(rows):
    """
    Escanea las primeras filas (lista de tuples) para detectar
    (header_idx, cols_dict, fmt) o (None, None, None).
    """
    for ri, row in enumerate(rows[:6]):
        texts = [str(v).lower().strip() for v in row if v]
        is_standard = any('posición del parque' in t or 'posicion del parque' in t for t in texts)
        is_oct2025  = any('nombre subestaci' in t for t in texts) and not is_standard

        if is_standard or is_oct2025:
            fmt = 'standard' if is_standard else 'oct2025'
            cols = {}
            for j, v in enumerate(row):
                if not v:
                    continue
                s  = str(v).strip()
                sl = s.lower()

                if fmt == 'standard':
                    if sl == 'provincia':
                        cols.setdefault('provincia', j)
                    if sl == 'nombre':
                        cols.setdefault('nombre', j)
                    if 'nivel de tens' in sl:
                        cols.setdefault('tension', j)
                    if 'posición del parque' in sl or 'posicion del parque' in sl:
                        cols['posicion'] = j
                    if 'capacidad disponible' in sl and 'firme' not in sl and 'prev' not in sl:
                        cols.setdefault('disponible', j)
                    if sl == 'fotovoltaico':
                        cols['fv'] = j
                    if 'eólico' in sl or 'eolico' in sl:
                        cols.setdefault('eolico', j)
                    if sl == 'otros':
                        cols.setdefault('otros', j)
                else:  # oct2025
                    if 'provincia' in sl and 'municipio' not in sl:
                        cols.setdefault('provincia', j)
                    if 'municipio' in sl:
                        cols.setdefault('municipio', j)
                    if 'nombre subestaci' in sl:
                        cols['nombre'] = j
                    if 'nivel de tens' in sl:
                        cols.setdefault('tension', j)
                    if sl == 'capacidad disponible (mw)':
                        cols['disponible'] = j
                    if 'admitida y no resuelta' in sl:
                        cols['en_estudio'] = j

            required = ('nombre', 'tension', 'posicion', 'disponible', 'fv', 'eolico', 'otros') \
                       if fmt == 'standard' else ('nombre', 'tension', 'disponible')
            if all(k in cols for k in required):
                return ri, cols, fmt
    return None, None, None


def process_standard(rows, fecha, header_idx, cols):
    groups = {}
    for row in rows[header_idx + 1:]:
        if not row or len(row) <= cols['nombre']:
            continue
        nombre = row[cols['nombre']]
        if not nombre or str(nombre).strip() in ('', 'Nombre'):
            continue
        nombre = str(nombre).strip()

        tension  = to_float(row[cols['tension']] if cols['tension'] < len(row) else None)
        posicion = str(row[cols['posicion']] if cols['posicion'] < len(row) else '').strip().lower()
        disp     = to_float(row[cols['disponible']] if cols['disponible'] < len(row) else None)
        fv       = to_float(row[cols['fv']]     if cols['fv']     < len(row) else None) or 0.0
        eol      = to_float(row[cols['eolico']] if cols['eolico'] < len(row) else None) or 0.0
        otros    = to_float(row[cols['otros']]  if cols['otros']  < len(row) else None) or 0.0
        prov     = str(row[cols['provincia']] if 'provincia' in cols and cols['provincia'] < len(row) else '').strip()

        key = (nombre, tension)
        if key not in groups:
            groups[key] = {'rows': [], 'total_parque': None, 'provincia': prov}
        if prov:
            groups[key]['provincia'] = prov

        entry = {'posicion': posicion, 'disp': disp, 'fv': fv, 'eol': eol, 'otros': otros}
        groups[key]['rows'].append(entry)
        if posicion == 'total parque':
            groups[key]['total_parque'] = entry

    records = []
    for (nombre, tension), gdata in groups.items():
        tp = gdata['total_parque'] or max(gdata['rows'], key=lambda r: r['disp'] or -1)
        disp = tp['disp']
        if disp is None:
            candidates = [r['disp'] for r in gdata['rows'] if r['disp'] is not None]
            disp = max(candidates) if candidates else 0.0
        disponible_neta = (disp or 0.0) - (tp['fv'] + tp['eol'] + tp['otros'])
        records.append({
            'fecha': fecha, 'distribuidora': 'UFD',
            'provincia': gdata['provincia'], 'municipio': '',
            'subestacion': nombre, 'tension_kv': tension,
            'disponible_neta': round(disponible_neta, 4),
        })
    return records


def process_oct2025(rows, fecha, header_idx, cols):
    # Verificar que hay datos de generación cacheados
    disp_idx = cols.get('disponible', -1)
    sample = [r[disp_idx] for r in rows[header_idx+1:header_idx+20] if r and disp_idx < len(r)]
    if all(v is None for v in sample):
        return []   # Fórmulas no cacheadas — sin datos de generación
    records = []
    for row in rows[header_idx + 1:]:
        if not row or len(row) <= cols['nombre']:
            continue
        nombre = row[cols['nombre']]
        if not nombre:
            continue
        nombre = str(nombre).strip()
        if not nombre or nombre.startswith('='):
            continue

        tension    = to_float(row[cols['tension']] if cols['tension'] < len(row) else None)
        disp       = to_float(row[cols['disponible']] if cols['disponible'] < len(row) else None)
        en_estudio = to_float(row[cols['en_estudio']] if 'en_estudio' in cols and cols['en_estudio'] < len(row) else None) or 0.0
        prov       = str(row[cols.get('provincia',-1)] if cols.get('provincia',-1) < len(row) and cols.get('provincia',-1) >= 0 else '').strip()
        mun        = str(row[cols.get('municipio',-1)]  if cols.get('municipio',-1)  < len(row) and cols.get('municipio',-1)  >= 0 else '').strip()

        # Si disponible es None (fórmulas no cacheadas) → tratar como 0
        disp = disp if disp is not None else 0.0
        records.append({
            'fecha': fecha, 'distribuidora': 'UFD',
            'provincia': prov, 'municipio': mun,
            'subestacion': nombre, 'tension_kv': tension,
            'disponible_neta': round(disp - en_estudio, 4),
        })
    return records


# ── main ──────────────────────────────────────────────────────────────────────

parser = argparse.ArgumentParser()
parser.add_argument('--summary', action='store_true')
args = parser.parse_args()

if not XLSX.exists():
    sys.exit(f"No se encuentra {XLSX}")

print(f"Leyendo {XLSX.name} …")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

all_records = []
skipped = []

for sname in wb.sheetnames:
    fecha = parse_sheet_date(sname)
    if not fecha:
        skipped.append((sname, 'nombre no parseado como fecha'))
        continue

    ws = wb[sname]
    # Leer TODAS las filas en memoria una sola vez (evita problemas de read_only)
    all_rows = list(ws.iter_rows(values_only=True))

    header_idx, cols, fmt = detect_format_and_cols(all_rows)
    if header_idx is None:
        skipped.append((sname, 'cabecera/columnas no detectadas'))
        continue

    if fmt == 'standard':
        records = process_standard(all_rows, fecha, header_idx, cols)
    else:
        records = process_oct2025(all_rows, fecha, header_idx, cols)

    all_records.extend(records)
    print(f"  {fecha}  ({sname!r:<26})  fmt={fmt:<9}  →  {len(records):>4} SETs")

wb.close()

if skipped:
    print(f"\n  Omitidas ({len(skipped)}):")
    for s, r in skipped:
        print(f"    {s!r}: {r}")

if not all_records:
    sys.exit("\nNo se encontraron registros.")

df = pd.DataFrame(all_records)
df['fecha'] = pd.to_datetime(df['fecha'])
df = df.sort_values(['subestacion', 'tension_kv', 'fecha'])

print(f"\n--- RESUMEN ---")
print(f"Total registros  : {len(df):,}")
print(f"Snapshots        : {df['fecha'].nunique()} "
      f"({df['fecha'].min().date()} → {df['fecha'].max().date()})")
print(f"SETs únicos      : {df['subestacion'].nunique():,}")
print(f"disp_neta (MW)   : min={df['disponible_neta'].min():.1f}  "
      f"max={df['disponible_neta'].max():.1f}  media={df['disponible_neta'].mean():.1f}")

if args.summary:
    sys.exit(0)

df['fecha'] = df['fecha'].dt.strftime('%Y-%m-%d')
df.to_csv(OUT, index=False, encoding='utf-8-sig')
print(f"Guardado en      : {OUT}")
